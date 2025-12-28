"""
مدیریت اجرت‌های استاندارد (تعرفه خدمات) - نسخه اصلاح شده و بهینه
با قابلیت‌های: کدگذاری خودکار، جلوگیری از تکرار، اعتبارسنجی و UI بهبود یافته
"""

import sys
import os
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QLineEdit, QComboBox, QTextEdit, QTableWidget, QCompleter,
    QTableWidgetItem, QMessageBox, QHeaderView, QFormLayout,
    QDoubleSpinBox, QSpinBox, QGroupBox, QScrollArea, QWidget,
    QFrame, QAbstractItemView, QSizePolicy
)
from PySide6.QtCore import Qt, Signal, QTimer
from PySide6.QtGui import QFont, QColor, QPalette


class ServiceFeeManager(QDialog):
    """
    فرم مدیریت اجرت‌های استاندارد - نسخه اصلاح شده
    """
    
    service_fee_updated = Signal()  # سیگنال به‌روزرسانی
    
    def __init__(self, data_manager):
        super().__init__()
        self.data_manager = data_manager
        self.current_service_id = None
        self.all_categories = []
        self.all_services = []
        
        self.setup_ui()
        self.load_initial_data()
        self.setup_stylesheet()
        
    def setup_ui(self):
        """راه‌اندازی رابط کاربری"""
        self.setWindowTitle("💰 مدیریت اجرت‌های استاندارد")
        self.setMinimumSize(900, 700)
        self.setMaximumSize(1200, 900)
        self.setLayoutDirection(Qt.RightToLeft)
        
        # ایجاد اسکرول‌اریا
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        content_widget = QWidget()
        main_layout = QVBoxLayout(content_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        # ========== بخش 1: فرم ورود اطلاعات ==========
        form_group = self.create_form_group()
        main_layout.addWidget(form_group)
        
        # ========== بخش 2: دکمه‌های عملیاتی ==========
        button_group = self.create_button_group()
        main_layout.addWidget(button_group)
        
        # ========== بخش 3: جدول لیست اجرت‌ها ==========
        table_group = self.create_table_group()
        main_layout.addWidget(table_group)
        
        # ========== بخش 4: دکمه‌های پایینی ==========
        bottom_group = self.create_bottom_group()
        main_layout.addWidget(bottom_group)
        
        main_layout.addStretch()
        
        scroll_area.setWidget(content_widget)
        
        # لایه اصلی
        dialog_layout = QVBoxLayout(self)
        dialog_layout.addWidget(scroll_area)
        
    def create_form_group(self):
        """ایجاد گروه فرم ورود اطلاعات"""
        group = QGroupBox("➕ افزودن/ویرایش اجرت")
        group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        layout = QFormLayout()
        layout.setSpacing(15)
        layout.setLabelAlignment(Qt.AlignRight)
        
        # کد خدمت
        self.service_code_input = QLineEdit()
        self.service_code_input.setPlaceholderText("به صورت خودکار تولید می‌شود")
        self.service_code_input.setReadOnly(True)
        self.service_code_input.setFixedHeight(35)
        
        # دکمه تولید کد جدید
        code_layout = QHBoxLayout()
        code_layout.addWidget(self.service_code_input)
        self.btn_generate_code = QPushButton("🔄 تولید کد")
        self.btn_generate_code.setFixedSize(100, 35)
        self.btn_generate_code.clicked.connect(self.generate_and_set_code)
        code_layout.addWidget(self.btn_generate_code)
        
        layout.addRow("کد خدمت:", code_layout)
        
        # نام خدمت
        self.service_name_input = QLineEdit()
        self.service_name_input.setPlaceholderText("مثال: تعویض کمپرسور یخچال")
        self.service_name_input.setFixedHeight(35)
        layout.addRow("* نام خدمت:", self.service_name_input)
        
        # دسته‌بندی
        category_layout = QHBoxLayout()
        category_layout.setSpacing(10)
        
        self.category_combo = QComboBox()
        self.category_combo.setEditable(True)
        self.category_combo.setInsertPolicy(QComboBox.NoInsert)
        self.category_combo.setCompleter(QCompleter([]))
        self.category_combo.setFixedHeight(35)
        self.category_combo.setMinimumWidth(250)
        self.category_combo.lineEdit().setPlaceholderText("جستجو یا وارد کردن دسته جدید...")
        
        self.btn_add_category = QPushButton("➕")
        self.btn_add_category.setFixedSize(35, 35)
        self.btn_add_category.setToolTip("افزودن دسته جدید")
        self.btn_add_category.clicked.connect(self.add_new_category)
        
        category_layout.addWidget(self.category_combo)
        category_layout.addWidget(self.btn_add_category)
        layout.addRow("* دسته‌بندی:", category_layout)
        
        # قیمت پیش‌فرض
        self.default_fee_input = QDoubleSpinBox()
        self.default_fee_input.setRange(0, 100000000)
        self.default_fee_input.setValue(0)
        self.default_fee_input.setPrefix("")
        self.default_fee_input.setSuffix(" تومان")
        self.default_fee_input.setButtonSymbols(QDoubleSpinBox.UpDownArrows)
        self.default_fee_input.setFixedHeight(35)
        self.default_fee_input.setMinimumWidth(200)
        layout.addRow("* قیمت پیش‌فرض:", self.default_fee_input)
        
        # ساعت تخمینی
        self.estimated_hours_input = QDoubleSpinBox()
        self.estimated_hours_input.setRange(0.1, 100.0)
        self.estimated_hours_input.setValue(1.0)
        self.estimated_hours_input.setSingleStep(0.5)
        self.estimated_hours_input.setDecimals(1)
        self.estimated_hours_input.setSuffix(" ساعت")
        self.estimated_hours_input.setFixedHeight(35)
        layout.addRow("ساعت تخمینی:", self.estimated_hours_input)
        
        # سطح سختی
        difficulty_layout = QHBoxLayout()
        self.difficulty_spin = QSpinBox()
        self.difficulty_spin.setRange(1, 5)
        self.difficulty_spin.setValue(1)
        self.difficulty_spin.setFixedHeight(35)
        difficulty_layout.addWidget(self.difficulty_spin)
        difficulty_layout.addWidget(QLabel("(۱=آسان، ۵=سخت)"))
        difficulty_layout.addStretch()
        layout.addRow("سطح سختی:", difficulty_layout)
        
        # توضیحات
        self.description_input = QTextEdit()
        self.description_input.setMaximumHeight(80)
        self.description_input.setPlaceholderText("توضیحات تکمیلی درباره این خدمت...")
        layout.addRow("توضیحات:", self.description_input)
        
        # وضعیت
        self.status_combo = QComboBox()
        self.status_combo.addItems(["فعال", "غیرفعال"])
        self.status_combo.setFixedHeight(35)
        layout.addRow("وضعیت:", self.status_combo)
        
        group.setLayout(layout)
        return group
    
    def create_button_group(self):
        """ایجاد گروه دکمه‌های عملیاتی"""
        frame = QFrame()
        layout = QHBoxLayout(frame)
        layout.setSpacing(15)
        
        self.btn_new = QPushButton("🆕 جدید")
        self.btn_new.setFixedHeight(40)
        self.btn_new.setStyleSheet(self.get_button_style("#27ae60"))
        self.btn_new.clicked.connect(self.new_service)
        
        self.btn_save = QPushButton("💾 ذخیره")
        self.btn_save.setFixedHeight(40)
        self.btn_save.setStyleSheet(self.get_button_style("#3498db"))
        self.btn_save.clicked.connect(self.save_service)
        
        self.btn_delete = QPushButton("🗑️ حذف")
        self.btn_delete.setFixedHeight(40)
        self.btn_delete.setStyleSheet(self.get_button_style("#e74c3c"))
        self.btn_delete.clicked.connect(self.delete_service)
        self.btn_delete.setEnabled(False)
        
        layout.addWidget(self.btn_new)
        layout.addWidget(self.btn_save)
        layout.addWidget(self.btn_delete)
        layout.addStretch()
        
        return frame
    
    def create_table_group(self):
        """ایجاد گروه جدول اجرت‌ها"""
        group = QGroupBox("📋 لیست اجرت‌های ثبت شده")
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # نوار جستجو
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 جستجو:"))
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("جستجو در نام خدمت، دسته‌بندی یا کد...")
        self.search_input.setFixedHeight(35)
        self.search_input.textChanged.connect(self.filter_table)
        search_layout.addWidget(self.search_input)
        
        self.btn_refresh = QPushButton("🔄")
        self.btn_refresh.setFixedSize(35, 35)
        self.btn_refresh.setToolTip("بروزرسانی لیست")
        self.btn_refresh.setStyleSheet(self.get_button_style("#3498db"))
        self.btn_refresh.clicked.connect(self.load_services)
        search_layout.addWidget(self.btn_refresh)
        
        layout.addLayout(search_layout)
        
        # جدول
        self.services_table = QTableWidget()
        self.services_table.setColumnCount(7)
        self.services_table.setHorizontalHeaderLabels([
            "کد خدمت", "نام خدمت", "دسته‌بندی", "قیمت (تومان)", 
            "ساعت تخمینی", "سطح سختی", "وضعیت"
        ])
        
        # تنظیمات جدول
        self.services_table.setAlternatingRowColors(True)
        self.services_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.services_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.services_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.services_table.setMinimumHeight(300)
        
        # تنظیم عرض ستون‌ها
        header = self.services_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)  # کد
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # نام
        header.setSectionResizeMode(2, QHeaderView.Fixed)  # دسته
        header.setSectionResizeMode(3, QHeaderView.Fixed)  # قیمت
        header.setSectionResizeMode(4, QHeaderView.Fixed)  # ساعت
        header.setSectionResizeMode(5, QHeaderView.Fixed)  # سطح
        header.setSectionResizeMode(6, QHeaderView.Fixed)  # وضعیت
        
        self.services_table.setColumnWidth(0, 120)
        self.services_table.setColumnWidth(2, 150)
        self.services_table.setColumnWidth(3, 130)
        self.services_table.setColumnWidth(4, 110)
        self.services_table.setColumnWidth(5, 90)
        self.services_table.setColumnWidth(6, 90)
        
        # اتصال رویداد انتخاب
        self.services_table.itemSelectionChanged.connect(self.on_service_selected)
        
        layout.addWidget(self.services_table)
        
        # آمار
        stats_layout = QHBoxLayout()
        self.stats_label = QLabel("0 اجرت ثبت شده")
        self.stats_label.setStyleSheet("font-weight: bold; color: #f39c12;")
        stats_layout.addWidget(QLabel("📊 آمار:"))
        stats_layout.addWidget(self.stats_label)
        stats_layout.addStretch()
        
        layout.addLayout(stats_layout)
        
        group.setLayout(layout)
        return group
    
    def create_bottom_group(self):
        """ایجاد گروه دکمه‌های پایینی"""
        frame = QFrame()
        layout = QHBoxLayout(frame)
        layout.setSpacing(20)
        
        self.btn_export = QPushButton("📤 خروجی Excel")
        self.btn_export.setFixedHeight(40)
        self.btn_export.setStyleSheet(self.get_button_style("#16a085"))
        self.btn_export.clicked.connect(self.export_to_excel)
        
        layout.addWidget(self.btn_export)
        layout.addStretch()
        
        self.btn_close = QPushButton("❌ بستن")
        self.btn_close.setFixedHeight(40)
        self.btn_close.setStyleSheet(self.get_button_style("#95a5a6"))
        self.btn_close.clicked.connect(self.close)
        
        layout.addWidget(self.btn_close)
        
        return frame
    
    def setup_stylesheet(self):
        """تنظیم استایل‌شیت"""
        self.setStyleSheet("""
            QDialog {
                background-color: #1e1e1e;
                color: white;
            }
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #3498db;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                background-color: #1e1e1e;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top right;
                padding: 5px 15px;
                background-color: #3498db;
                color: white;
                border-radius: 4px;
            }
            QLineEdit, QComboBox, QTextEdit, QDoubleSpinBox, QSpinBox {
                background-color: #2c2c2c;
                color: white;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 6px;
                font-size: 12px;
            }
            QLineEdit:focus, QComboBox:focus, QTextEdit:focus {
                border: 2px solid #3498db;
            }
            QTableWidget {
                background-color: #2c2c2c;
                alternate-background-color: #3a3a3a;
                gridline-color: #444;
                border: 1px solid #555;
                border-radius: 5px;
            }
            QTableWidget::item {
                padding: 6px;
                border-bottom: 1px solid #444;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QHeaderView::section {
                background-color: #2c3e50;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
        """)
    
    def get_button_style(self, color):
        """استایل دکمه"""
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
                border: none;
                min-width: 80px;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
            QPushButton:disabled {{
                background-color: #7f8c8d;
                color: #bdc3c7;
            }}
        """
    
    def darken_color(self, color, amount=30):
        """تیره کردن رنگ"""
        color = color.lstrip('#')
        r, g, b = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
        r, g, b = max(0, r-amount), max(0, g-amount), max(0, b-amount)
        return f'#{r:02x}{g:02x}{b:02x}'
    
    def load_initial_data(self):
        """بارگذاری داده‌های اولیه"""
        self.load_categories()
        self.load_services()
        self.new_service()  # شروع با فرم خالی
    
    def load_categories(self):
        """بارگذاری دسته‌بندی‌ها"""
        try:
            # دریافت دسته‌بندی‌های منحصر به فرد از دیتابیس
            query = """
            SELECT DISTINCT category 
            FROM ServiceFees 
            WHERE category IS NOT NULL AND category != '' 
            ORDER BY category
            """
            categories = self.data_manager.service_fee.fetch_all(query)
            
            self.all_categories = [cat['category'] for cat in categories]
            
            # اضافه کردن دسته‌های پیش‌فرض اگر خالی بود
            if not self.all_categories:
                self.all_categories = [
                    "عمومی", "یخچال", "کولر گازی", "ماشین لباسشویی",
                    "آبگرمکن", "ماشین ظرفشویی", "اجاق گاز", "هود",
                    "جاروبرقی", "پنکه", "سایر"
                ]
            
            # به‌روزرسانی ComboBox
            self.category_combo.clear()
            self.category_combo.addItems(self.all_categories)
            
            # تنظیم تکمیل‌کننده
            completer = QCompleter(self.all_categories)
            completer.setCaseSensitivity(Qt.CaseInsensitive)
            completer.setFilterMode(Qt.MatchContains)
            self.category_combo.setCompleter(completer)
            
        except Exception as e:
            print(f"خطا در بارگذاری دسته‌بندی‌ها: {e}")
            self.all_categories = []
    
    def add_new_category(self):
        """افزودن دسته‌بندی جدید"""
        category = self.category_combo.currentText().strip()
        
        if not category:
            QMessageBox.warning(self, "خطا", "لطفا نام دسته‌بندی را وارد کنید.")
            return
        
        if category in self.all_categories:
            QMessageBox.information(self, "تکراری", 
                                  f"دسته‌بندی '{category}' قبلاً وجود دارد.")
            return
        
        # افزودن به لیست
        self.all_categories.append(category)
        self.all_categories.sort()
        
        # به‌روزرسانی ComboBox
        self.category_combo.clear()
        self.category_combo.addItems(self.all_categories)
        
        # انتخاب دسته جدید
        index = self.category_combo.findText(category)
        if index >= 0:
            self.category_combo.setCurrentIndex(index)
        
        QMessageBox.information(self, "موفق", 
                              f"دسته‌بندی '{category}' با موفقیت اضافه شد.")
    
    def generate_service_code(self):
        """تولید کد خدمت خودکار"""
        try:
            # دریافت آخرین کد
            query = """
            SELECT service_code 
            FROM ServiceFees 
            WHERE service_code LIKE 'SRV-%' 
            ORDER BY id DESC 
            LIMIT 1
            """
            result = self.data_manager.service_fee.fetch_one(query)
            
            if result and result['service_code']:
                # استخراج عدد آخر
                last_code = result['service_code']
                try:
                    last_num = int(last_code.split('-')[1])
                    new_num = last_num + 1
                except (IndexError, ValueError):
                    new_num = 1
            else:
                new_num = 1
            
            # تولید کد جدید
            new_code = f"SRV-{new_num:04d}"
            
            # بررسی تکراری نبودن
            counter = 1
            while self.check_duplicate_code(new_code):
                new_num += 1
                new_code = f"SRV-{new_num:04d}"
                counter += 1
                if counter > 1000:  # جلوگیری از حلقه بی‌نهایت
                    import time
                    timestamp = int(time.time() % 10000)
                    new_code = f"SRV-{timestamp:04d}"
                    break
            
            return new_code
            
        except Exception as e:
            print(f"خطا در تولید کد: {e}")
            import time
            timestamp = int(time.time() % 10000)
            return f"SRV-{timestamp:04d}"
    
    def check_duplicate_code(self, service_code, exclude_id=None):
        """بررسی تکراری بودن کد خدمت"""
        try:
            if exclude_id:
                query = """
                SELECT id 
                FROM ServiceFees 
                WHERE service_code = ? AND id != ?
                """
                params = (service_code, exclude_id)
            else:
                query = "SELECT id FROM ServiceFees WHERE service_code = ?"
                params = (service_code,)
            
            result = self.data_manager.service_fee.fetch_one(query, params)
            return result is not None
            
        except Exception as e:
            print(f"خطا در بررسی تکراری بودن کد: {e}")
            return True  # در صورت خطا، فرض بر تکراری بودن
    
    def check_duplicate_service(self, service_name, category, exclude_id=None):
        """بررسی تکراری بودن نام خدمت در دسته"""
        try:
            if exclude_id:
                query = """
                SELECT id 
                FROM ServiceFees 
                WHERE service_name = ? AND category = ? AND id != ?
                """
                params = (service_name, category, exclude_id)
            else:
                query = """
                SELECT id 
                FROM ServiceFees 
                WHERE service_name = ? AND category = ?
                """
                params = (service_name, category)
            
            result = self.data_manager.service_fee.fetch_one(query, params)
            return result is not None
            
        except Exception as e:
            print(f"خطا در بررسی تکراری بودن خدمت: {e}")
            return True
    
    def generate_and_set_code(self):
        """تولید و تنظیم کد جدید"""
        new_code = self.generate_service_code()
        self.service_code_input.setText(new_code)
    
    def load_services(self):
        """بارگذاری لیست اجرت‌ها"""
        try:
            # دریافت تمام خدمات
            self.all_services = self.data_manager.service_fee.get_all_services(
                active_only=False
            )
            
            # پاک کردن جدول
            self.services_table.setRowCount(0)
            
            # پر کردن جدول
            for service in self.all_services:
                row = self.services_table.rowCount()
                self.services_table.insertRow(row)
                
                # ذخیره ID در UserRole
                item_code = QTableWidgetItem(service['service_code'])
                item_code.setData(Qt.UserRole, service['id'])
                
                self.services_table.setItem(row, 0, item_code)
                self.services_table.setItem(row, 1, 
                    QTableWidgetItem(service['service_name']))
                self.services_table.setItem(row, 2, 
                    QTableWidgetItem(service['category']))
                self.services_table.setItem(row, 3, 
                    QTableWidgetItem(f"{service['default_fee']:,}"))
                self.services_table.setItem(row, 4, 
                    QTableWidgetItem(str(service['estimated_hours'] or '1.0')))
                self.services_table.setItem(row, 5, 
                    QTableWidgetItem(str(service['difficulty_level'] or '1')))
                
                # وضعیت
                status_text = "فعال" if service['is_active'] else "غیرفعال"
                status_item = QTableWidgetItem(status_text)
                if service['is_active']:
                    status_item.setForeground(QColor('#27ae60'))
                    status_item.setFont(QFont('', weight=QFont.Bold))
                else:
                    status_item.setForeground(QColor('#e74c3c'))
                self.services_table.setItem(row, 6, status_item)
            
            # به‌روزرسانی آمار
            active_count = sum(1 for s in self.all_services if s['is_active'])
            self.stats_label.setText(
                f"{len(self.all_services)} اجرت ثبت شده ({active_count} فعال)"
            )
            
        except Exception as e:
            print(f"خطا در بارگذاری اجرت‌ها: {e}")
            QMessageBox.critical(self, "خطا", 
                               f"خطا در بارگذاری لیست اجرت‌ها:\n{str(e)}")
    
    def filter_table(self):
        """فیلتر کردن جدول بر اساس جستجو"""
        search_text = self.search_input.text().lower().strip()
        
        if not search_text:
            # نمایش همه سطرها
            for row in range(self.services_table.rowCount()):
                self.services_table.setRowHidden(row, False)
            return
        
        for row in range(self.services_table.rowCount()):
            match = False
            
            # جستجو در کد، نام و دسته
            for col in [0, 1, 2]:
                item = self.services_table.item(row, col)
                if item and search_text in item.text().lower():
                    match = True
                    break
            
            self.services_table.setRowHidden(row, not match)
    
    def on_service_selected(self):
        """هنگام انتخاب سطر از جدول"""
        selected = self.services_table.selectedItems()
        if not selected:
            self.btn_delete.setEnabled(False)
            return
        
        row = selected[0].row()
        service_id = self.services_table.item(row, 0).data(Qt.UserRole)
        
        # دریافت اطلاعات خدمت
        service = self.data_manager.service_fee.fetch_one(
            "SELECT * FROM ServiceFees WHERE id = ?", 
            (service_id,)
        )
        
        if service:
            self.current_service_id = service['id']
            self.service_code_input.setText(service['service_code'])
            self.service_name_input.setText(service['service_name'])
            
            # تنظیم دسته‌بندی
            cat_index = self.category_combo.findText(service['category'])
            if cat_index >= 0:
                self.category_combo.setCurrentIndex(cat_index)
            else:
                self.category_combo.setCurrentText(service['category'])
            
            self.default_fee_input.setValue(service['default_fee'])
            self.estimated_hours_input.setValue(service['estimated_hours'] or 1.0)
            self.difficulty_spin.setValue(service['difficulty_level'] or 1)
            self.description_input.setText(service.get('description', ''))
            self.status_combo.setCurrentText(
                "فعال" if service['is_active'] else "غیرفعال"
            )
            
            self.btn_delete.setEnabled(True)
    
    def validate_form(self):
        """اعتبارسنجی فرم"""
        errors = []
        
        # نام خدمت
        service_name = self.service_name_input.text().strip()
        if not service_name:
            errors.append("نام خدمت الزامی است")
        
        # دسته‌بندی
        category = self.category_combo.currentText().strip()
        if not category:
            errors.append("دسته‌بندی الزامی است")
        
        # قیمت
        if self.default_fee_input.value() <= 0:
            errors.append("قیمت باید بیشتر از صفر باشد")
        
        # کد خدمت
        service_code = self.service_code_input.text().strip()
        if not service_code:
            errors.append("کد خدمت الزامی است")
        elif self.check_duplicate_code(service_code, self.current_service_id):
            errors.append(f"کد خدمت '{service_code}' تکراری است")
        
        # نام تکراری در دسته
        if service_name and category:
            if self.check_duplicate_service(
                service_name, category, self.current_service_id
            ):
                errors.append(
                    f"خدمت '{service_name}' در دسته '{category}' قبلاً ثبت شده است"
                )
        
        return errors
    
    def new_service(self):
        """ایجاد خدمت جدید"""
        self.current_service_id = None
        self.service_code_input.clear()
        self.generate_and_set_code()  # تولید کد جدید
        self.service_name_input.clear()
        self.category_combo.setCurrentIndex(-1)
        self.category_combo.clearEditText()
        self.default_fee_input.setValue(0)
        self.estimated_hours_input.setValue(1.0)
        self.difficulty_spin.setValue(1)
        self.description_input.clear()
        self.status_combo.setCurrentText("فعال")
        
        # غیرفعال کردن دکمه حذف
        self.btn_delete.setEnabled(False)
        
        # لغو انتخاب در جدول
        self.services_table.clearSelection()
        
        # فوکوس روی نام خدمت
        self.service_name_input.setFocus()
    
    def save_service(self):
        """ذخیره یا به‌روزرسانی اجرت"""
        # اعتبارسنجی
        errors = self.validate_form()
        if errors:
            QMessageBox.warning(
                self, 
                "خطا در اعتبارسنجی", 
                "لطفا خطاهای زیر را برطرف کنید:\n\n• " + "\n• ".join(errors)
            )
            return
        
        try:
            # جمع‌آوری داده‌ها
            service_data = {
                'service_code': self.service_code_input.text().strip(),
                'service_name': self.service_name_input.text().strip(),
                'category': self.category_combo.currentText().strip(),
                'default_fee': self.default_fee_input.value(),
                'estimated_hours': self.estimated_hours_input.value(),
                'difficulty_level': self.difficulty_spin.value(),
                'description': self.description_input.toPlainText(),
                'is_active': 1 if self.status_combo.currentText() == "فعال" else 0
            }
            
            if self.current_service_id:
                # ویرایش
                success = self.data_manager.service_fee.update_service(
                    self.current_service_id, service_data
                )
                message = "ویرایش"
            else:
                # افزودن جدید
                success = self.data_manager.service_fee.add_service(service_data)
                message = "ثبت"
            
            if success:
                QMessageBox.information(
                    self, "موفق", 
                    f"اجرت با موفقیت {message} شد."
                )
                
                # تازه‌سازی
                self.load_categories()  # ممکن است دسته جدید اضافه شده باشد
                self.load_services()
                self.service_fee_updated.emit()
                
                if not self.current_service_id:
                    self.new_service()  # آماده برای ورودی بعدی
                    
            else:
                QMessageBox.critical(
                    self, "خطا", 
                    f"خطا در {message} اجرت."
                )
                
        except Exception as e:
            QMessageBox.critical(
                self, "خطا", 
                f"خطا در ذخیره اجرت:\n{str(e)}"
            )
            print(f"خطای کامل: {e}")
    
    def delete_service(self):
        """حذف اجرت"""
        if not self.current_service_id:
            QMessageBox.warning(self, "خطا", "لطفا ابتدا یک اجرت را انتخاب کنید.")
            return
        
        # دریافت اطلاعات برای نمایش
        service = self.data_manager.service_fee.fetch_one(
            "SELECT service_code, service_name, category FROM ServiceFees WHERE id = ?",
            (self.current_service_id,)
        )
        
        if not service:
            QMessageBox.warning(self, "خطا", "خدمت مورد نظر یافت نشد.")
            return
        
        reply = QMessageBox.question(
            self, "تأیید حذف",
            f"آیا مطمئن هستید که می‌خواهید این اجرت را حذف کنید؟\n\n"
            f"• کد: {service['service_code']}\n"
            f"• نام: {service['service_name']}\n"
            f"• دسته: {service['category']}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                success = self.data_manager.service_fee.delete_service(
                    self.current_service_id
                )
                
                if success:
                    QMessageBox.information(self, "موفق", "اجرت با موفقیت حذف شد.")
                    self.load_services()
                    self.service_fee_updated.emit()
                    self.new_service()  # بازنشانی فرم
                else:
                    QMessageBox.critical(self, "خطا", "خطا در حذف اجرت.")
                    
            except Exception as e:
                QMessageBox.critical(
                    self, "خطا", 
                    f"خطا در حذف اجرت:\n{str(e)}"
                )
    
    def export_to_excel(self):
        """خروجی Excel"""
        try:
            # بررسی وجود openpyxl
            try:
                from openpyxl import Workbook
                from datetime import datetime
            except ImportError:
                QMessageBox.warning(
                    self, "نصب وابستگی",
                    "برای خروجی Excel نیاز به نصب openpyxl است:\n\n"
                    "pip install openpyxl\n\n"
                    "یا:\n"
                    "pip install -r requirements.txt"
                )
                return
            
            # دریافت خدمات فعال
            services = self.data_manager.service_fee.get_active_services()
            if not services:
                QMessageBox.information(self, "خالی", "هیچ خدمت فعالی برای خروجی وجود ندارد.")
                return
            
            # ایجاد Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "اجرت‌های استاندارد"
            
            # هدرها
            headers = [
                "کد خدمت", "نام خدمت", "دسته‌بندی", 
                "قیمت (تومان)", "ساعت تخمینی", 
                "سطح سختی", "توضیحات"
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                ws.cell(row=1, column=col).font = Workbook().create_font(bold=True)
            
            # داده‌ها
            for row, service in enumerate(services, 2):
                ws.cell(row=row, column=1, value=service['service_code'])
                ws.cell(row=row, column=2, value=service['service_name'])
                ws.cell(row=row, column=3, value=service['category'])
                ws.cell(row=row, column=4, value=service['default_fee'])
                ws.cell(row=row, column=5, value=service.get('estimated_hours', 1.0))
                ws.cell(row=row, column=6, value=service.get('difficulty_level', 1))
                ws.cell(row=row, column=7, value=service.get('description', ''))
            
            # تنظیم عرض ستون‌ها
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # ذخیره فایل
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"اجرت_های_استاندارد_{timestamp}.xlsx"
            wb.save(filename)
            
            QMessageBox.information(
                self, "موفق",
                f"فایل Excel با موفقیت ایجاد شد:\n\n{filename}\n\n"
                f"تعداد رکوردها: {len(services)}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self, "خطا",
                f"خطا در ایجاد فایل Excel:\n{str(e)}"
            )